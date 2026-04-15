#!/usr/bin/env python3
"""Compare our bug analysis with the previous Excel analysis."""

import sys
from openpyxl import load_workbook
from collections import Counter

def read_excel_analysis(excel_path):
    """Read the previous Excel analysis."""
    print("=" * 60)
    print("Reading Previous Excel Analysis")
    print("=" * 60)

    wb = load_workbook(excel_path, data_only=True)

    print(f"\n📊 Worksheets found: {wb.sheetnames}")

    # Read the "Testable Bugs" sheet (main analysis)
    if "Testable Bugs" in wb.sheetnames:
        sheet = wb["Testable Bugs"]
    else:
        sheet = wb.active

    print(f"\n📋 Reading sheet: {sheet.title}")

    # Get headers
    headers = []
    for cell in sheet[1]:
        if cell.value:
            headers.append(cell.value)

    print(f"\n📌 Headers: {headers}")

    # Read all rows
    bugs = []
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):  # Skip empty rows
            continue

        bug_data = dict(zip(headers, row))
        if bug_data.get('Key') or bug_data.get('Issue Key') or bug_data.get('Jira Key'):
            bugs.append(bug_data)

    print(f"\n✅ Total bugs in Excel: {len(bugs)}")

    # Print first few rows to understand structure
    if bugs:
        print("\n📋 Sample bug data:")
        for i, bug in enumerate(bugs[:3], 1):
            print(f"\n{i}. Bug data:")
            for key, value in bug.items():
                if value is not None:
                    print(f"   {key}: {value}")

    return bugs, headers


def analyze_excel_data(bugs, headers):
    """Analyze the Excel data to extract statistics."""
    print("\n" + "=" * 60)
    print("Excel Analysis Statistics")
    print("=" * 60)

    stats = {
        'total': len(bugs),
        'coverage': Counter(),
        'test_level': Counter(),
        'categories': Counter(),
        'priorities': Counter()
    }

    # Try to identify which columns contain what data
    coverage_col = None
    test_level_col = None
    category_col = None
    priority_col = None

    # Common column name variations
    for header in headers:
        header_lower = str(header).lower()

        if 'coverage' in header_lower or 'status' in header_lower:
            coverage_col = header
        elif 'test level' in header_lower or 'level' in header_lower:
            test_level_col = header
        elif 'categor' in header_lower or 'type' in header_lower:
            category_col = header
        elif 'priority' in header_lower:
            priority_col = header

    print(f"\n🔍 Identified columns:")
    print(f"   Coverage: {coverage_col}")
    print(f"   Test Level: {test_level_col}")
    print(f"   Category: {category_col}")
    print(f"   Priority: {priority_col}")

    # Collect statistics
    for bug in bugs:
        if coverage_col and bug.get(coverage_col):
            stats['coverage'][str(bug[coverage_col])] += 1

        if test_level_col and bug.get(test_level_col):
            stats['test_level'][str(bug[test_level_col])] += 1

        if category_col and bug.get(category_col):
            stats['categories'][str(bug[category_col])] += 1

        if priority_col and bug.get(priority_col):
            stats['priorities'][str(bug[priority_col])] += 1

    # Print statistics
    print(f"\n📊 Coverage Breakdown ({coverage_col}):")
    for status, count in stats['coverage'].most_common():
        pct = int(count / stats['total'] * 100)
        print(f"   {status:25} {count:3} ({pct:2}%)")

    print(f"\n🎯 Test Level Breakdown ({test_level_col}):")
    for level, count in stats['test_level'].most_common():
        pct = int(count / stats['total'] * 100)
        print(f"   {level:25} {count:3} ({pct:2}%)")

    print(f"\n🏷️  Category Breakdown ({category_col}):")
    for cat, count in stats['categories'].most_common():
        pct = int(count / stats['total'] * 100)
        print(f"   {cat:25} {count:3} ({pct:2}%)")

    print(f"\n⚡ Priority Breakdown ({priority_col}):")
    for priority, count in stats['priorities'].most_common():
        pct = int(count / stats['total'] * 100)
        print(f"   {priority:25} {count:3} ({pct:2}%)")

    return stats


def compare_with_our_analysis(excel_stats):
    """Compare Excel analysis with our automated analysis."""
    print("\n" + "=" * 60)
    print("Comparison: Excel vs Our Automated Analysis")
    print("=" * 60)

    # Our analysis results (from the full test run)
    our_stats = {
        'total': 249,
        'coverage': {
            'COVERED': 186,
            'PARTIALLY COVERED': 51,
            'GAP': 1,
            'NOT TESTABLE': 11
        },
        'test_level': {
            'Mock': 194,
            'E2E': 42,
            'Unit': 4,
            'N/A': 9
        },
        'categories': {
            'functional': 223,
            'upgrade': 14,
            'security': 6,
            'performance': 6,
            'fips': 1,
            'disconnected': 1
        }
    }

    print(f"\n📊 Total Bugs:")
    print(f"   Excel:     {excel_stats['total']}")
    print(f"   Our Tool:  {our_stats['total']}")

    delta = our_stats['total'] - excel_stats['total']
    if delta > 0:
        print(f"   → Our tool found {delta} more bugs (possibly newer bugs)")
    elif delta < 0:
        print(f"   → Excel has {abs(delta)} more bugs (may include resolved/older)")
    else:
        print(f"   ✅ Exact match!")

    print("\n" + "=" * 60)
    print("Key Differences & Insights")
    print("=" * 60)

    return our_stats


def main():
    # Accept Excel path as command-line argument
    if len(sys.argv) < 2:
        print("Usage: compare_analyses.py <excel_file_path>")
        print("Example: compare_analyses.py ./bug-analysis.xlsx")
        sys.exit(1)

    excel_path = sys.argv[1]

    print("\n🔍 Historical Bug Coverage Analysis - Comparison Tool\n")

    try:
        # Read Excel
        excel_bugs, headers = read_excel_analysis(excel_path)

        # Analyze Excel data
        excel_stats = analyze_excel_data(excel_bugs, headers)

        # Compare
        our_stats = compare_with_our_analysis(excel_stats)

        print("\n✅ Comparison complete!")

        print("\n💡 Recommendations:")
        print("   1. Review any bugs in Excel not in our analysis (may be older/resolved)")
        print("   2. Review any bugs in our analysis not in Excel (newer bugs)")
        print("   3. Compare classification methodology differences")
        print("   4. Validate test level assignments for any discrepancies")

        return 0

    except Exception as e:
        print(f"\n❌ Error: {e}")
        import traceback
        traceback.print_exc()
        return 1


if __name__ == "__main__":
    sys.exit(main())
